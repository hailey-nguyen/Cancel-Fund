import yfinance as yf
import pandas as pd
import openpyxl

# ================================================
# 1. Input Section
# ================================================
# List of stock symbols to analyze
stock_symbols = ['COIN', 'AAL', 'LUV', 'UAL', 'GM', 'VLKAF', 'BYDDY']  

# Corresponding RSI Scores for each stock
# Must be in the same order as the stock_symbols list
# Change these values manually as needed
rsi_scores = [35.5, 30.75, 37.43, 38.21, 35.42, 55.75, 81.67]

# Corresponding Twitter Post Scores and Sentiment Scores for each stock
# Must be in the same order as the stock_symbols list
twitter_post_scores = [4, 8, 8, 8, 4, 4, 4]  # Directly use these values (0–10 scale)
sent_scores = [-0.118, -0.121, -0.121, -0.121, -0.0554, -0.0554, -0.0554]

# Weighting Constants
RSI_WEIGHT = 0.2
ANALYST_WEIGHT = 0.2
TWITTER_WEIGHT = 0.3  
SENTIMENT_WEIGHT = 0.3  # Renamed from NEWS_WEIGHT
HOLDINGS_NUM = 5

# ================================================
# 2. Scoring Functions
# ================================================

def analyst_score(ar: str):
    # Transform analyst ratings into a score of 10
    ar = ar.lower()
    if ar == "strong_buy":
        return 10
    elif ar == "buy":
        return 7.5
    elif ar == "neutral":
        return 5
    elif ar == "hold":
        return 5
    elif ar == "sell":
        return 2.5
    elif ar == "strong_sell":
        return 0
    elif ar == "underperform":
        return 0
    return 0

def rsi_score(rsi: float):
    # Transform RSI scores into a score of 10
    if rsi <= 30:
        return 10
    elif 30 < rsi <= 40:
        return 8
    elif 40 < rsi <= 50:
        return 6
    elif 50 < rsi <= 60:
        return 4
    elif 60 < rsi <= 70:
        return 2
    else:
        return 0

def sentiment_score(sentiment: float):
    # Inverted scoring: More negative sentiment gets a higher score
    if sentiment < -0.5:
        return 10  # Very Negative
    elif -0.5 <= sentiment < -0.4:
        return 8
    elif -0.4 <= sentiment < -0.3:
        return 7
    elif -0.3 <= sentiment < -0.2:
        return 6
    elif -0.2 <= sentiment < -0.1:
        return 5
    elif -0.1 <= sentiment < 0:
        return 4
    elif 0 <= sentiment < 0.1:
        return 3
    elif 0.1 <= sentiment < 0.2:
        return 2
    else:  # sentiment >= 0.2
        return 0

# ================================================
# 3. Get Data from Yahoo Finance using yfinance
# ================================================
def get_yfinance_data(stock_symbols):
    data = []

    for symbol in stock_symbols:
        print(f"Collecting data for {symbol}...")

        # Get Stock Data from Yahoo Finance
        stock = yf.Ticker(symbol)

        # Manually Input RSI from the list
        rsi_index = stock_symbols.index(symbol)
        rsi = rsi_scores[rsi_index]

        # Get Analyst Rating
        try:
            analyst = stock.info['recommendationKey'].capitalize() if 'recommendationKey' in stock.info else "Neutral"
        except:
            analyst = "Neutral"
            print(f"Analyst Rating not found for {symbol}")

        data.append({
            'Symbol': symbol,
            'RSI': rsi,
            'Analyst Rating': analyst
        })

    return pd.DataFrame(data)

# ================================================
# 4. Main Script: Pull Data and Calculate Scores
# ================================================

# Get Data from Yahoo Finance
df = get_yfinance_data(stock_symbols)

# Calculate Scores
portfolio_data = []

for i, row in df.iterrows():
    symbol = row['Symbol']
    rsi = row['RSI']
    analyst = row['Analyst Rating']

    rsi_score_value = rsi_score(rsi) if rsi else 0
    analyst_score_value = analyst_score(analyst)

    # Use Twitter Post Scores as-is
    twitter_post_score = twitter_post_scores[i]  # Directly use the value provided

    # Sentiment Scores using sentiment_score()
    sentiment_score_value = sentiment_score(sent_scores[i])

    # Calculate Total Score
    total_score = (rsi_score_value * RSI_WEIGHT +
                   analyst_score_value * ANALYST_WEIGHT +
                   twitter_post_score * TWITTER_WEIGHT +
                   sentiment_score_value * SENTIMENT_WEIGHT)

    portfolio_data.append([symbol, rsi, analyst, rsi_score_value, analyst_score_value, 
                           twitter_post_score, sentiment_score_value, total_score])

# Convert to DataFrame
df = pd.DataFrame(portfolio_data, columns=[
    'Symbol', 'RSI', 'Analyst Rating', 'RSI Score', 'Analyst Score', 
    'Twitter Post Score', 'Sentiment Score', 'Total Score'])

# Sort and Select Top Holdings
df = df.sort_values(by='Total Score', ascending=False)
portfolio = df.head(HOLDINGS_NUM)

# Weighting the Portfolio
total_of_scores = portfolio['Total Score'].sum()
portfolio['Weight'] = portfolio['Total Score'] / total_of_scores

# Display Portfolio
print("\nSelected Portfolio:")
print(portfolio)

# ================================================
# 5. Overwrite Existing Excel File
# ================================================
# ================================================
# 5. Overwrite Existing Excel File with Text Wrapping
# ================================================
output_file = "fund_holdings.xlsx"
with pd.ExcelWriter(output_file, mode='w', engine='openpyxl') as writer:
    portfolio.to_excel(writer, sheet_name="Portfolio", index=False)

# Apply Text Wrapping to All Cells
workbook = openpyxl.load_workbook(output_file)
worksheet = workbook['Portfolio']

# Loop through all cells and apply text wrapping
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

# Save the Workbook
workbook.save(output_file)

print(f"\nPortfolio saved to {output_file} (overwritten with text wrapping).")


# ================================================
# 6. Summary and Output
# ================================================
print(f"\nPortfolio saved to {output_file} (overwritten).")
