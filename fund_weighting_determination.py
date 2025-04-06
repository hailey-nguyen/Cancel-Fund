import yfinance as yf
import pandas as pd
import openpyxl
import os

# ================================================
# 1. Input Section
# ================================================
# These are the stock tickers to evaluate. The last three are designated shorts.
stock_symbols = ['NFLX', 'CMCSA', 'ILMN', 'VCYT', 'GM', 'VLKAF', 'DIS', 'TSLA', 'ME']

# Manually input RSI scores for each ticker (must be in the same order as stock_symbols)
rsi_scores = [65.06, 56.84, 39.65, 55.38, 53.96, 47.62, 68.23, 74.75, 19.09]

# Manually input Twitter post activity scores (0–10 scale)
twitter_post_scores = [4, 4, 6, 6, 4, 4, 4, 4, 6]

# Manually input normalized sentiment scores (negative = bearish sentiment)
sent_scores = [0.06286, 0.06286, -0.009, -0.009, 0.00119, 0.00119, 0.06286, 0.00119, -0.009]

# These are the weights for each factor contributing to the final score
RSI_WEIGHT = 0.2           # 20% of total score from RSI
ANALYST_WEIGHT = 0.2       # 20% from analyst rating
TWITTER_WEIGHT = 0.3       # 30% from Twitter post volume
SENTIMENT_WEIGHT = 0.3     # 30% from sentiment score

HOLDINGS_NUM = 4               # Top 4 scoring long positions will be selected
TOTAL_PORTFOLIO_ALLOCATION = 80  # Portfolio is 80% invested (20% held in cash)

# ================================================
# 2. Scoring Functions
# ================================================

# Converts analyst ratings into a 10-point scale
def analyst_score(ar: str):
    ar = ar.lower()
    return {
        "strong_buy": 10,
        "buy": 7.5,
        "neutral": 5,
        "hold": 5,
        "sell": 2.5,
        "strong_sell": 0,
        "underperform": 0
    }.get(ar, 0)

# Converts RSI into a 10-point scale (lower RSI is more attractive for mean reversion)
def rsi_score(rsi: float):
    if rsi <= 30: return 10
    elif rsi <= 40: return 8
    elif rsi <= 50: return 6
    elif rsi <= 60: return 4
    elif rsi <= 70: return 2
    else: return 0

# Converts sentiment score into a 10-point scale (more negative sentiment = higher score)
def sentiment_score(sentiment: float):
    if sentiment < -0.5: return 10
    elif sentiment < -0.4: return 8
    elif sentiment < -0.3: return 7
    elif sentiment < -0.2: return 6
    elif sentiment < -0.1: return 5
    elif sentiment < 0: return 4
    elif sentiment < 0.1: return 3
    elif sentiment < 0.2: return 2
    else: return 0

# ================================================
# 3. Get Data from Yahoo Finance
# ================================================
def get_yfinance_data(stock_symbols):
    data = []
    for symbol in stock_symbols:
        print(f"Collecting data for {symbol}...")

        # Fetch the stock object
        stock = yf.Ticker(symbol)

        # Retrieve RSI from manual input
        rsi_index = stock_symbols.index(symbol)
        rsi = rsi_scores[rsi_index]

        # Retrieve analyst rating
        try:
            analyst = stock.info['recommendationKey'].capitalize() if 'recommendationKey' in stock.info else "Neutral"
        except:
            analyst = "Neutral"

        data.append({'Symbol': symbol, 'RSI': rsi, 'Analyst Rating': analyst})
    return pd.DataFrame(data)

# ================================================
# 4. Calculate Scores and Create DataFrame
# ================================================
df = get_yfinance_data(stock_symbols)

portfolio_data = []

for i, row in df.iterrows():
    symbol = row['Symbol']
    rsi = row['RSI']
    analyst = row['Analyst Rating']

    # Get individual factor scores
    rsi_val = rsi_score(rsi)
    analyst_val = analyst_score(analyst)
    twitter_val = twitter_post_scores[i]
    sentiment_val = sentiment_score(sent_scores[i])

    # Calculate total score using a weighted sum:
    # Total_Score = RSI*0.2 + Analyst*0.2 + Twitter*0.3 + Sentiment*0.3
    total_score = (
        rsi_val * RSI_WEIGHT +
        analyst_val * ANALYST_WEIGHT +
        twitter_val * TWITTER_WEIGHT +
        sentiment_val * SENTIMENT_WEIGHT
    )

    portfolio_data.append([
        symbol, rsi, analyst, rsi_val, analyst_val, 
        twitter_val, sentiment_val, total_score
    ])

# Create a complete DataFrame from all calculated scores
df = pd.DataFrame(portfolio_data, columns=[
    'Symbol', 'RSI', 'Analyst Rating', 'RSI Score', 'Analyst Score',
    'Twitter Post Score', 'Sentiment Score', 'Total Score'
])

# ================================================
# 5. Split Long and Short Positions
# ================================================
# These 3 symbols will be shorted
short_symbols = ['DIS', 'TSLA', 'ME']

# Exclude shorts and rank remaining stocks by total score to select longs
long_candidates = df[~df['Symbol'].isin(short_symbols)]
long_stocks = long_candidates.nlargest(HOLDINGS_NUM, 'Total Score')

# Include only the designated short positions
short_stocks = df[df['Symbol'].isin(short_symbols)]

# Compute combined total score of both longs and shorts
total_scores = long_stocks['Total Score'].sum() + short_stocks['Total Score'].sum()

# Assign weights as percentages of total portfolio (80% cap)
# Formula: Weight_i = 0.80 * (Score_i / Σ Score_j)
long_stocks['Weight'] = (long_stocks['Total Score'] / total_scores) * TOTAL_PORTFOLIO_ALLOCATION
short_stocks['Weight'] = (short_stocks['Total Score'] / total_scores) * TOTAL_PORTFOLIO_ALLOCATION * -1  # Display negative weight for shorts

# ================================================
# 6. Save to Excel: One Sheet with Two Tables
# ================================================
output_file = "fund_holdings.xlsx"

# Remove corrupt file if it exists
if os.path.exists(output_file):
    try:
        openpyxl.load_workbook(output_file)
    except:
        print("\nCorrupted Excel file detected. Deleting and recreating.")
        os.remove(output_file)

# Write long positions to Excel
with pd.ExcelWriter(output_file, mode='w', engine='openpyxl') as writer:
    long_stocks.to_excel(writer, sheet_name="Portfolio", index=False, startrow=0)

# Open workbook and prepare to insert a blank row between tables
workbook = openpyxl.load_workbook(output_file)
worksheet = workbook["Portfolio"]
separator_row = worksheet.max_row + 2
for col in range(1, worksheet.max_column + 1):
    worksheet.cell(row=separator_row, column=col, value="")  # Insert blank row
workbook.save(output_file)

# Append short positions below the separator
with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists="overlay") as writer:
    short_stocks.to_excel(writer, sheet_name="Portfolio", index=False, startrow=separator_row)

print(f"\nPortfolio saved to {output_file} (long and short positions included).")

# ================================================
# 7. Display in Console
# ================================================
print("\nLong Positions:")
print(long_stocks)
print("\nShort Positions:")
print(short_stocks)
